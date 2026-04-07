import os
import smtplib
import re
import time
import base64
import io
from email.message import EmailMessage
from openpyxl import load_workbook

def execute_email(config, input_data, context):
    subject_template = config.get("subject", "")
    body_template = config.get("body", "")
    manual_emails = config.get("manual_emails", [])
    excel_file = config.get("excel_file")

    # --- Configuration Validation ---
    if not subject_template or not body_template:
        return {
            "status": "error",
            "node": "email",
            "message": "Configuration error",
            "error": "Subject or Body template is missing",
            "hint": "Please provide both subject and body templates."
        }

    recipients = []

    # --- Recipient Gathering ---
    if excel_file:
        try:
            # Decode base64 excel data
            file_content = base64.b64decode(excel_file['data'])
            wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
            sheet = wb.active
            
            # Skip header, assume Col 1: Email, Col 2: Name
            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                email, name = row
                if email and re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", str(email)):
                    recipients.append({"email": str(email).strip(), "name": str(name or "").strip()})
        except Exception as e:
            return {
                "status": "error",
                "node": "email",
                "message": "Excel parsing failed",
                "error": str(e),
                "hint": "Ensure the file is a valid .xlsx and follow the column instructions."
            }
    else:
        # Use manual emails
        for email in manual_emails:
            if email and re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", str(email)):
                recipients.append({"email": str(email).strip(), "name": "User"})

    if not recipients:
        return {
            "status": "error",
            "node": "email",
            "message": "No valid recipients",
            "error": "Recipient list is empty",
            "hint": "Add manual emails or upload a valid Excel file."
        }

    # --- Limit Check ---
    if len(recipients) > 500:
        return {
            "status": "error",
            "node": "email",
            "message": "Limit exceeded",
            "error": f"Found {len(recipients)} recipients",
            "hint": "Maximum 500 emails allowed per run."
        }

    # --- SMTP Auth ---
    email_user = os.environ.get("EMAIL_USER", "").strip()
    email_pass = os.environ.get("EMAIL_PASS", "").strip().replace(" ", "")

    if not email_user or not email_pass:
        return {
            "status": "error",
            "node": "email",
            "message": "SMTP Configuration missing",
            "error": "EMAIL_USER or EMAIL_PASS not set",
            "hint": "Check your .env file credentials."
        }

    # --- Bulk Sending with Batching ---
    total = len(recipients)
    sent = 0
    failed = 0
    batch_size = 20
    delay_between_batches = 2

    try:
        # Open a single persistent connection for the whole run
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(email_user, email_pass)
            
            for i, recipient in enumerate(recipients):
                # Batch delay
                if i > 0 and i % batch_size == 0:
                    time.sleep(delay_between_batches)

                try:
                    # Personalize content
                    subject = subject_template.replace("{{name}}", recipient["name"])
                    body = body_template.replace("{{name}}", recipient["name"])

                    msg = EmailMessage()
                    msg.set_content(body)
                    msg["Subject"] = subject
                    msg["From"] = email_user
                    msg["To"] = recipient["email"]

                    smtp.send_message(msg)
                    sent += 1
                except Exception:
                    failed += 1

        return {
            "status": "success",
            "node": "email",
            "message": f"Bulk email complete: {sent} sent, {failed} failed",
            "data": {
                "total": total,
                "sent": sent,
                "failed": failed
            }
        }

    except smtplib.SMTPAuthenticationError:
        return {
            "status": "error",
            "node": "email",
            "message": "SMTP Auth failed",
            "error": "Invalid credentials",
            "hint": "Verify your Gmail App Password."
        }
    except Exception as e:
        return {
            "status": "error",
            "node": "email",
            "message": "Bulk send aborted",
            "error": str(e),
            "hint": "Check network connection or SMTP limits."
        }
